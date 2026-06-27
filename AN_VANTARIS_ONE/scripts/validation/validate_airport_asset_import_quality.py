from pathlib import Path
from collections import Counter, defaultdict
import hashlib
import json
from datetime import datetime, timezone
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]

INPUTS = [
    ROOT / "AN_VANTARIS_ONE" / "source-data" / "Asset Database Zonewise_Basement_SH_030626.xlsx",
    ROOT / "AN_VANTARIS_ONE" / "source-data" / "Asset Database Zonewise_Ground Floor_SH_220626.xlsx",
]

REPORT_DIR = ROOT / "AN_VANTARIS_ONE" / "reports" / "asset-import-ga"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

REQUIRED_COLUMNS = [
    "Device ID", "Building", "Level", "Zone", "System", "Location", "Device Type",
]

ALLOWED_SYSTEMS = {"PA", "ACS", "CCTV", "TEL", "IPTV", "RAS", "MCS"}

LOCATION_HINTS = {
    "BASEMET": "BASEMENT",
    "MARKUP": "MAKEUP",
    "MEETERS-GREETERS": "MEETERS GREETERS",
    "DEPT": "DEPARTURE",
    " RM": " ROOM",
    "&": "AND",
}

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def norm(v):
    return "" if v is None else str(v).strip()

def sha256_file(path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def issue(severity, category, message, count, sample=None):
    return {
        "severity": severity,
        "category": category,
        "message": message,
        "count": int(count),
        "sample": sample or [],
    }

def read_workbook(path):
    if not path.exists():
        raise FileNotFoundError(f"Missing source file: {path}")

    file_hash = sha256_file(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            continue

        headers = [norm(x) for x in values[0]]
        if not any(headers):
            continue

        for row_idx, row in enumerate(values[1:], start=2):
            record = {}
            for i, header in enumerate(headers):
                if header:
                    record[header] = norm(row[i] if i < len(row) else "")

            if not any(record.values()):
                continue

            record["__source_file"] = path.name
            record["__source_hash"] = file_hash
            record["__sheet"] = sheet_name
            record["__row"] = row_idx
            rows.append(record)

    return rows

def sample_rows(rows, limit=10):
    return [
        {
            "source_file": r.get("__source_file", ""),
            "sheet": r.get("__sheet", ""),
            "row": r.get("__row", ""),
            "SL": r.get("SL", ""),
            "Device ID": r.get("Device ID", ""),
            "Building": r.get("Building", ""),
            "Level": r.get("Level", ""),
            "Zone": r.get("Zone", ""),
            "System": r.get("System", ""),
            "Location": r.get("Location", ""),
            "Device Type": r.get("Device Type", ""),
        }
        for r in rows[:limit]
    ]

records = []
source_file_profiles = []

for path in INPUTS:
    rows = read_workbook(path)
    records.extend(rows)
    source_file_profiles.append({
        "file_name": path.name,
        "file_hash_sha256": sha256_file(path),
        "records": len(rows),
    })

if not records:
    raise SystemExit("BLOCKER: no records parsed")

all_columns = set()
for r in records:
    all_columns.update(r.keys())

issues = []

missing_columns = [c for c in REQUIRED_COLUMNS if c not in all_columns]
if missing_columns:
    issues.append(issue(
        "BLOCKER", "schema",
        f"Required columns missing: {missing_columns}",
        len(missing_columns),
        [{"missing_column": c} for c in missing_columns],
    ))

for col in REQUIRED_COLUMNS:
    missing = [r for r in records if r.get(col, "") == ""]
    if missing:
        severity = "BLOCKER" if col in ["Device ID", "Building", "Level", "Zone", "System", "Device Type"] else "MAJOR"
        issues.append(issue(
            severity, "required_field",
            f"Missing required field: {col}",
            len(missing),
            sample_rows(missing)
        ))

device_groups = defaultdict(list)
for r in records:
    did = r.get("Device ID", "")
    if did:
        device_groups[did].append(r)

duplicate_ids = {did: rows for did, rows in device_groups.items() if len(rows) > 1}
if duplicate_ids:
    dup_sample = []
    for did, rows in list(duplicate_ids.items())[:10]:
        for r in rows[:3]:
            dup_sample.append({
                "Device ID": did,
                "source_file": r.get("__source_file", ""),
                "sheet": r.get("__sheet", ""),
                "row": r.get("__row", ""),
                "System": r.get("System", ""),
                "Location": r.get("Location", ""),
                "Device Type": r.get("Device Type", ""),
            })
    issues.append(issue(
        "BLOCKER", "duplicate_device_id",
        "Duplicate Device ID detected",
        len(duplicate_ids),
        dup_sample
    ))

    conflict_count = 0
    conflict_samples = []
    for did, rows in duplicate_ids.items():
        conflicts = {}
        for field in ["System", "Location", "Device Type", "Level", "Zone"]:
            vals = sorted(set(r.get(field, "") for r in rows))
            if len(vals) > 1:
                conflicts[field] = vals
        if conflicts:
            conflict_count += 1
            if len(conflict_samples) < 10:
                conflict_samples.append({"Device ID": did, "conflicts": conflicts})
    if conflict_count:
        issues.append(issue(
            "BLOCKER", "duplicate_conflict",
            "Duplicate Device ID has conflicting attributes",
            conflict_count,
            conflict_samples
        ))

unknown_systems = [r for r in records if r.get("System", "") and r.get("System", "") not in ALLOWED_SYSTEMS]
if unknown_systems:
    issues.append(issue(
        "MAJOR", "unknown_system",
        "System value is not in approved dictionary",
        len(unknown_systems),
        sample_rows(unknown_systems)
    ))

for col in ["Last Done", "Due Date", "Status", "Overdue"]:
    if col in all_columns:
        missing = [r for r in records if r.get(col, "") == ""]
        if missing:
            issues.append(issue(
                "WARNING", "maintenance_field",
                f"Missing maintenance field: {col}",
                len(missing),
                sample_rows(missing, 5)
            ))

for col in ["Area", "Remarks"]:
    if col in all_columns:
        missing = [r for r in records if r.get(col, "") == ""]
        if missing:
            issues.append(issue(
                "WARNING", "optional_field",
                f"Empty optional field: {col}",
                len(missing),
                []
            ))

locations = sorted(set(r.get("Location", "") for r in records if r.get("Location", "")))
hinted = []
for loc in locations:
    suggested = loc
    for old, new in LOCATION_HINTS.items():
        suggested = suggested.replace(old, new)
    if suggested != loc:
        hinted.append({"location": loc, "suggested": suggested})

if hinted:
    issues.append(issue(
        "WARNING", "location_normalization",
        "Location naming normalization suggestions detected",
        len(hinted),
        hinted[:30]
    ))

severity_counts = Counter(i["severity"] for i in issues)

readiness = "READY"
if severity_counts["BLOCKER"] > 0:
    readiness = "HOLD_BLOCKED"
elif severity_counts["MAJOR"] > 0:
    readiness = "REVIEW_REQUIRED"
elif severity_counts["WARNING"] > 0:
    readiness = "READY_WITH_WARNINGS"

confirm_enabled = readiness != "HOLD_BLOCKED"
requires_second_confirmation = severity_counts["WARNING"] > 0 or severity_counts["MAJOR"] > 0

summary = {
    "total_records": len(records),
    "source_files": sorted(set(r.get("__source_file", "") for r in records)),
    "sheet_count": len(sorted(set(r.get("__sheet", "") for r in records))),
    "level_count": len(sorted(set(r.get("Level", "") for r in records if r.get("Level", "")))),
    "zone_count": len(sorted(set(r.get("Zone", "") for r in records if r.get("Zone", "")))),
    "system_count": len(sorted(set(r.get("System", "") for r in records if r.get("System", "")))),
    "device_type_count": len(sorted(set(r.get("Device Type", "") for r in records if r.get("Device Type", "")))),
    "location_count": len(locations),
    "blocker_count": severity_counts.get("BLOCKER", 0),
    "major_count": severity_counts.get("MAJOR", 0),
    "warning_count": severity_counts.get("WARNING", 0),
    "info_count": severity_counts.get("INFO", 0),
    "readiness": readiness,
}

alert = {
    "title": "Asset Import Quality Check",
    "message": "Customer asset workbook has been parsed and validated. Review the quality report before import.",
    "readiness": readiness,
    "confirm_enabled": confirm_enabled,
    "requires_review": readiness in ["HOLD_BLOCKED", "REVIEW_REQUIRED", "READY_WITH_WARNINGS"],
    "requires_second_confirmation": requires_second_confirmation,
    "action_labels": [
        "Cancel Import",
        "Download Quality Report",
        "Review Issues",
        "Confirm Import",
    ],
}

payload = {
    "report_id": "ONE_ASSET_IMPORT_GA_R1_QUALITY_REPORT",
    "generated_at": now_iso(),
    "source_file_profiles": source_file_profiles,
    "summary": summary,
    "alert": alert,
    "system_summary": Counter(r.get("System", "") for r in records).most_common(),
    "level_summary": Counter(r.get("Level", "") for r in records).most_common(),
    "zone_summary": Counter(r.get("Zone", "") for r in records).most_common(),
    "device_type_summary": Counter(r.get("Device Type", "") for r in records).most_common(80),
    "location_summary": Counter(r.get("Location", "") for r in records).most_common(100),
    "issues": issues,
}

json_path = REPORT_DIR / "airport_asset_import_quality_report.json"
json_path.write_text(json.dumps(payload, indent=2))

md = []
md.append("# Airport Asset Import Data Quality Report\n")
md.append("## Production Import Readiness\n")
md.append(f"- Readiness: **{readiness}**")
md.append(f"- Confirm enabled: {confirm_enabled}")
md.append(f"- Requires second confirmation: {requires_second_confirmation}\n")

md.append("## Summary\n")
for k, v in summary.items():
    md.append(f"- {k}: {v}")

md.append("\n## Source Files\n")
for f in source_file_profiles:
    md.append(f"- {f['file_name']}: {f['records']} records, sha256={f['file_hash_sha256']}")

md.append("\n## Issue Summary\n")
for sev in ["BLOCKER", "MAJOR", "WARNING", "INFO"]:
    md.append(f"- {sev}: {severity_counts.get(sev, 0)}")

md.append("\n## Import Alert Payload\n")
md.append("```json")
md.append(json.dumps(alert, indent=2))
md.append("```")

md.append("\n## Issues\n")
for i in issues:
    md.append(f"### {i['severity']} / {i['category']}")
    md.append(f"- {i['message']}")
    md.append(f"- Count: {i['count']}")
    if i["sample"]:
        md.append("- Sample:")
        md.append("```json")
        md.append(json.dumps(i["sample"], indent=2))
        md.append("```")

md.append("\n## Top Systems\n")
for name, count in Counter(r.get("System", "") for r in records).most_common():
    md.append(f"- {name or '(blank)'}: {count}")

md.append("\n## Top Device Types\n")
for name, count in Counter(r.get("Device Type", "") for r in records).most_common(40):
    md.append(f"- {name or '(blank)'}: {count}")

md.append("\n## Top Locations\n")
for name, count in Counter(r.get("Location", "") for r in records).most_common(40):
    md.append(f"- {name or '(blank)'}: {count}")

md_path = REPORT_DIR / "airport_asset_import_quality_report.md"
md_path.write_text("\n".join(md))

print("=== ASSET IMPORT QUALITY VALIDATION ===")
print(json.dumps(summary, indent=2))
print("=== ALERT ===")
print(json.dumps(alert, indent=2))
print(f"WROTE {json_path}")
print(f"WROTE {md_path}")

if readiness == "HOLD_BLOCKED":
    raise SystemExit(2)
