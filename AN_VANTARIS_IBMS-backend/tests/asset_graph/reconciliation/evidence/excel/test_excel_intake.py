"""Tests for synthetic Excel evidence intake adapter."""
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[6]
BACKEND = ROOT / "AN_VANTARIS_IBMS-backend"
sys.path.insert(0, str(BACKEND))

from src.asset_graph.reconciliation.evidence.excel.converter import convert_workbook_rows
from src.asset_graph.reconciliation.evidence.excel.errors import ExcelIntakeError
from src.asset_graph.reconciliation.evidence.excel.intake import (
    compare_deterministic_outputs,
    run_excel_evidence_intake,
)
from src.asset_graph.reconciliation.evidence.excel.validator import (
    validate_required_columns,
    validate_workbook_profile,
)
from src.asset_graph.reconciliation.evidence.excel.workbook import ExcelWorkbook
import importlib.util

FIXTURES_PATH = Path(__file__).resolve().parent / "fixtures.py"
_spec = importlib.util.spec_from_file_location("excel_fixtures", FIXTURES_PATH)
_fixtures = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_fixtures)
write_workbook = _fixtures.write_workbook

EXCEL_RUNNER = ROOT / "scripts/validation/run-one-device-reconciliation-excel.py"
MODULE_DIR = BACKEND / "src/asset_graph/reconciliation/evidence/excel"


class TestExcelEvidenceIntake(unittest.TestCase):
    maxDiff = None

    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.workbook_path = Path(self.tempdir.name) / "fixture.xlsx"
        write_workbook(self.workbook_path)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    # 1
    def test_workbook_extension_validation_rejects_xlsm(self) -> None:
        path = Path(self.tempdir.name) / "bad.xlsm"
        path.write_bytes(b"fake")
        with self.assertRaises(ExcelIntakeError):
            ExcelWorkbook.validate_extension(path)

    # 2
    def test_workbook_extension_accepts_xlsx(self) -> None:
        ExcelWorkbook.validate_extension(self.workbook_path)

    # 3
    def test_required_worksheet_validation(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            for sheet in ("Summary", "Devices", "StandardFields", "ScenarioCatalog", "RejectedSamples", "README"):
                self.assertIn(sheet, book.sheet_names())

    # 4
    def test_missing_worksheet_rejected(self) -> None:
        bad_path = Path(self.tempdir.name) / "missing-sheet.xlsx"
        with zipfile.ZipFile(self.workbook_path, "r") as zin:
            names = [name for name in zin.namelist() if "worksheets/sheet" in name]
            with zipfile.ZipFile(bad_path, "w") as zout:
                for item in zin.infolist():
                    if item.filename.endswith("workbook.xml"):
                        data = zin.read(item.filename).decode("utf-8").replace('name="Devices"', 'name="MissingDevices"')
                        zout.writestr(item, data.encode("utf-8"))
                    else:
                        zout.writestr(item, zin.read(item.filename))
        with self.assertRaises(ExcelIntakeError):
            ExcelWorkbook.open(bad_path)

    # 5
    def test_required_column_validation(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            rows = book.rows("Devices")
            validate_required_columns(tuple(rows[0]), ("RowID", "SourceID", "DeviceName"), "Devices")

    # 6
    def test_duplicate_headers_rejected(self) -> None:
        bad_path = Path(self.tempdir.name) / "dup-header.xlsx"
        write_workbook(bad_path)
        # rebuild with duplicate header via openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(bad_path)
        ws = wb["Devices"]
        ws["B1"] = "RowID"
        wb.save(bad_path)
        with self.assertRaises(ExcelIntakeError):
            with ExcelWorkbook.open(bad_path) as book:
                book.rows("Devices")

    # 7
    def test_empty_devices_rejected(self) -> None:
        with self.assertRaises(ExcelIntakeError):
            validate_workbook_profile([], [{"ScenarioClass": "CLEAN"}], [], enforce_fixture_profile=False)

    # 8
    def test_device_mapping(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            package = convert_workbook_rows(book.rows("Devices")[:1], [])
        device = package["devices"][0]
        self.assertEqual(device["sourceId"], "SYNTH-DEV-CLEAN-001")
        self.assertEqual(device["name"], "Synthetic Controller 001")

    # 9
    def test_standard_field_mapping(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            package = convert_workbook_rows([], book.rows("StandardFields")[:1])
        field = package["standardFields"][0]
        self.assertIn("sourceId", field)
        self.assertIn("deviceSourceId", field)

    # 10
    def test_empty_source_id_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            devices = book.rows("Devices")
        missing = [row for row in devices if row.get("Scenario") == "MISSING_SOURCE_ID"][0]
        package = convert_workbook_rows([missing], [])
        self.assertNotIn("sourceId", package["devices"][0])

    # 11
    def test_missing_device_name_preserved(self) -> None:
        row = {"ScenarioClass": "BLOCKER", "TenantID": "SYNTH-TENANT-001", "SiteID": "SYNTH-SITE-001",
               "SourceNamespace": "legacy.synth.excel.devices.ns1", "SourceID": "SYNTH-DEV-NONAME",
               "DeviceName": "", "DeviceType": "CONTROLLER", "OperationalStatus": "AVAILABLE"}
        package = convert_workbook_rows([row], [])
        self.assertNotIn("name", package["devices"][0])

    # 12
    def test_duplicate_source_identity_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            devices = [row for row in book.rows("Devices") if row.get("SourceID") == "SYNTH-DEV-DUP-001"]
        package = convert_workbook_rows(devices, [])
        self.assertEqual(len(package["devices"]), 2)

    # 13
    def test_tenant_mismatch_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            row = [item for item in book.rows("Devices") if item.get("Scenario") == "TENANT_MISMATCH"][0]
        package = convert_workbook_rows([row], [])
        self.assertEqual(package["devices"][0]["tenantId"], "SYNTH-TENANT-MISMATCH")

    # 14
    def test_site_mismatch_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            row = [item for item in book.rows("Devices") if item.get("Scenario") == "SITE_MISMATCH"][0]
        package = convert_workbook_rows([row], [])
        self.assertEqual(package["devices"][0]["siteId"], "SYNTH-SITE-MISMATCH")

    # 15
    def test_point_unknown_device_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            row = [item for item in book.rows("StandardFields") if item.get("Scenario") == "ORPHAN_POINT"][0]
        package = convert_workbook_rows([], [row])
        self.assertEqual(package["standardFields"][0]["deviceSourceId"], "SYNTH-DEV-NOT-EXISTS")

    # 16
    def test_unknown_status_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            row = [item for item in book.rows("Devices") if item.get("Scenario") == "UNKNOWN_STATUS"][0]
        package = convert_workbook_rows([row], [])
        self.assertEqual(package["devices"][0]["operationalStatus"], "UNKNOWN")

    # 17
    def test_missing_unit_preserved(self) -> None:
        row = {"ScenarioClass": "REVIEW", "SourceID": "F1", "DeviceSourceID": "D1", "FieldName": "x",
               "DisplayName": "X", "FieldType": "float", "DataType": "NUMBER", "Unit": "", "AccessMode": "READ_ONLY"}
        package = convert_workbook_rows([], [row])
        self.assertEqual(package["standardFields"][0]["unit"], "")

    # 18
    def test_missing_datatype_preserved(self) -> None:
        row = {"ScenarioClass": "CLEAN", "SourceID": "F2", "DeviceSourceID": "D1", "FieldName": "y",
               "DisplayName": "Y", "FieldType": "float", "DataType": "", "AccessMode": "READ_ONLY"}
        package = convert_workbook_rows([], [row])
        self.assertEqual(package["standardFields"][0]["dataType"], "STRING")

    # 19
    def test_tag_namespace_conflict_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            rows = [item for item in book.rows("Devices") if item.get("Scenario") == "TAG_COLLISION"]
        package = convert_workbook_rows(rows, [])
        tags = [item["sourceTagName"] for item in package["devices"]]
        self.assertTrue(all(tag == "SYN-TAG-COLLISION" for tag in tags))

    # 20
    def test_unresolved_location_preserved(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            row = [item for item in book.rows("Devices") if item.get("Scenario") == "UNRESOLVED_PARENT"][0]
        package = convert_workbook_rows([row], [])
        self.assertIn("parentSourceId", package["devices"][0]["locationReference"])

    # 21
    def test_accepted_sheet_privacy_scan_passes_fixture(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            profile = validate_workbook_profile(book.rows("Devices"), book.rows("StandardFields"), book.rows("RejectedSamples"))
        self.assertEqual(profile["deviceRows"], 60)

    # 22
    def test_rejected_samples_isolated(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            package = convert_workbook_rows(book.rows("Devices"), book.rows("StandardFields"))
        text = json.dumps(package)
        self.assertNotIn("SYNTHETIC_PASSWORD_SHOULD_BE_REJECTED", text)

    # 23
    def test_credential_rejection(self) -> None:
        intake = run_excel_evidence_intake(
            root=ROOT,
            input_path=self.workbook_path,
            json_output=Path(self.tempdir.name) / "pkg.json",
            report_output=Path(self.tempdir.name) / "report.json",
            run_id="TEST-CRED-001",
        )
        categories = {item["category"] for item in intake["rejectedSampleResults"]}
        self.assertIn("CREDENTIAL_FIELD", categories)

    # 24
    def test_telemetry_rejection(self) -> None:
        intake = run_excel_evidence_intake(
            root=ROOT,
            input_path=self.workbook_path,
            json_output=Path(self.tempdir.name) / "pkg2.json",
            report_output=Path(self.tempdir.name) / "report2.json",
            run_id="TEST-TEL-001",
        )
        statuses = {item["rejectionStatus"] for item in intake["rejectedSampleResults"]}
        self.assertEqual(statuses, {"REJECTED"})

    # 25
    def test_no_formulas_in_business_sheets(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            book.rows("Devices")
            book.rows("StandardFields")

    # 26
    def test_no_external_links(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            self.assertFalse(book.has_external_links())

    # 27
    def test_deterministic_json_package(self) -> None:
        out1 = Path(self.tempdir.name) / "p1.json"
        out2 = Path(self.tempdir.name) / "p2.json"
        run_excel_evidence_intake(root=ROOT, input_path=self.workbook_path, json_output=out1, report_output=Path(self.tempdir.name) / "r1.json", run_id="DET-001")
        run_excel_evidence_intake(root=ROOT, input_path=self.workbook_path, json_output=out2, report_output=Path(self.tempdir.name) / "r2.json", run_id="DET-001")
        self.assertEqual(out1.read_bytes(), out2.read_bytes())

    # 28
    def test_deterministic_reconciliation_report(self) -> None:
        package = Path(self.tempdir.name) / "same-package.json"
        r1 = Path(self.tempdir.name) / "dr1.json"
        r2 = Path(self.tempdir.name) / "dr2.json"
        run_excel_evidence_intake(root=ROOT, input_path=self.workbook_path, json_output=package, report_output=r1, run_id="DET-002")
        run_excel_evidence_intake(root=ROOT, input_path=self.workbook_path, json_output=package, report_output=r2, run_id="DET-002")
        ok, reason = compare_deterministic_outputs(r1, r2)
        self.assertTrue(ok, reason)

    # 29
    def test_no_database_imports(self) -> None:
        text = "\n".join(path.read_text(encoding="utf-8") for path in MODULE_DIR.rglob("*.py"))
        self.assertNotIn("sqlalchemy", text.lower())

    # 30
    def test_no_provider_writes(self) -> None:
        text = "\n".join(path.read_text(encoding="utf-8") for path in MODULE_DIR.rglob("*.py"))
        self.assertNotIn("db.session", text)

    # 31
    def test_no_public_api(self) -> None:
        text = "\n".join(path.read_text(encoding="utf-8") for path in MODULE_DIR.rglob("*.py"))
        self.assertNotIn("@api_bp.route", text)

    # 32
    def test_no_write_cutover_approval(self) -> None:
        report = Path(self.tempdir.name) / "report-wc.json"
        intake = run_excel_evidence_intake(
            root=ROOT,
            input_path=self.workbook_path,
            json_output=Path(self.tempdir.name) / "pkg-wc.json",
            report_output=report,
            run_id="TEST-WC-001",
        )
        self.assertEqual(intake["writeCutoverStatus"], "NOT_READY_FOR_WRITE_CUTOVER")
        payload = json.loads(report.read_text(encoding="utf-8"))
        self.assertNotEqual(payload.get("cutoverDecision"), "READY_FOR_WRITE_CUTOVER")

    # 33
    def test_generated_package_contains_assessment_type(self) -> None:
        out = Path(self.tempdir.name) / "assessment-type.json"
        run_excel_evidence_intake(root=ROOT, input_path=self.workbook_path, json_output=out, report_output=Path(self.tempdir.name) / "r.json", run_id="AT-001")
        package = json.loads(out.read_text(encoding="utf-8"))
        self.assertEqual(package["assessmentType"], "SYNTHETIC_REPRESENTATIVE_ONLY")

    # 34
    def test_fixture_profile_counts(self) -> None:
        with ExcelWorkbook.open(self.workbook_path) as book:
            profile = validate_workbook_profile(book.rows("Devices"), book.rows("StandardFields"), book.rows("RejectedSamples"))
        self.assertEqual(profile["deviceScenarioDistribution"]["CLEAN"], 48)
        self.assertEqual(profile["fieldScenarioDistribution"]["BLOCKER"], 60)

    # 35
    def test_cli_runner_passes(self) -> None:
        out_json = Path(self.tempdir.name) / "cli-package.json"
        out_report = Path(self.tempdir.name) / "cli-report.json"
        result = subprocess.run(
            [
                sys.executable, str(EXCEL_RUNNER), "--root", str(ROOT),
                "--input", str(self.workbook_path),
                "--json-output", str(out_json),
                "--report-output", str(out_report),
                "--run-id", "CLI-001",
            ],
            cwd=ROOT,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertIn("ONE_DEVICE_RECONCILIATION_EXCEL_RUN_PASS", result.stdout)

    # 36
    def test_blockers_visible_in_report(self) -> None:
        report = Path(self.tempdir.name) / "blocker-report.json"
        run_excel_evidence_intake(
            root=ROOT,
            input_path=self.workbook_path,
            json_output=Path(self.tempdir.name) / "blocker-package.json",
            report_output=report,
            run_id="BLOCKER-001",
        )
        payload = json.loads(report.read_text(encoding="utf-8"))
        self.assertTrue(payload.get("blockers"))

    # 37
    def test_reviews_visible_in_report(self) -> None:
        report = Path(self.tempdir.name) / "review-report.json"
        run_excel_evidence_intake(
            root=ROOT,
            input_path=self.workbook_path,
            json_output=Path(self.tempdir.name) / "review-package.json",
            report_output=report,
            run_id="REVIEW-001",
        )
        payload = json.loads(report.read_text(encoding="utf-8"))
        self.assertTrue(payload.get("reviews"))


if __name__ == "__main__":
    unittest.main()
