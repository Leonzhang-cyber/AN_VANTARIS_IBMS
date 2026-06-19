#!/usr/bin/env python3
"""Formula workbook fixtures for airport intake tests."""
from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

_FIXTURES_PATH = Path(__file__).resolve().parent / "fixtures.py"
_spec = importlib.util.spec_from_file_location("airport_fixtures", _FIXTURES_PATH)
_fixtures = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_fixtures)
HEADERS = _fixtures.HEADERS
_row = _fixtures._row


def write_formula_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws1 = wb.create_sheet("Zone-1")
    ws1.append(HEADERS)
    ws1.append(_row("1", "TE3-CCT-BAS-DA21-FCT-001", system="CCT"))
    ws1["A2"] = '=IF(B2<>"",1,"")'
    ws1["C2"] = '=LEFT(B2,3)'
    ws1["D2"] = '=IF(ISNUMBER(FIND("-BAS-",B2)),"BAS","GRD")'
    ws1["E2"] = ArrayFormula("E2:E2", "Z1")
    ws1["F2"] = '=MID(B2,12,4)'
    ws1["G2"] = '=IF(ISNUMBER(FIND("-CCT-",B2)),"CCTV","ACS")'
    ws1["J2"] = '=IF(ISNUMBER(FIND("-FCT-",B2)),"Fixed Camera","Unknown")'
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-TEL-BAS-DA31-TEL-001", system="TEL", da="DA31", level="BAS", zone="Z2"))
    ws2["K2"] = '=UPPER(I2)'
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_unapproved_formula_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws1 = wb.create_sheet("Zone-1")
    ws1.append(HEADERS)
    ws1.append(_row("1", "TE3-CCT-BAS-DA21-FCT-001"))
    ws1["I2"] = '=UPPER(H2)'
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-CCT-GRD-DA21-FCT-002", level="GRD", zone="Z2"))
    wb.save(path)


def write_external_formula_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws1 = wb.create_sheet("Zone-1")
    ws1.append(HEADERS)
    ws1.append(_row("1", "TE3-CCT-BAS-DA21-FCT-001"))
    ws1["C2"] = '=INDIRECT("[Other.xlsx]Sheet1!A1")'
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-CCT-GRD-DA21-FCT-002", level="GRD", zone="Z2"))
    wb.save(path)


def write_mismatch_formula_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws1 = wb.create_sheet("Zone-1")
    ws1.append(HEADERS)
    base = _row("1", "TE3-CCT-BAS-DA21-FCT-001", system="CCT")
    ws1.append(base)
    ws1["C2"] = "WRONG"
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-CCT-GRD-DA21-FCT-002", level="GRD", zone="Z2"))
    wb.save(path)
