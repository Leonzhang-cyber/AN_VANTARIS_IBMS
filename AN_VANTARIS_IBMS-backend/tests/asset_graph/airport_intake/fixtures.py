#!/usr/bin/env python3
"""Synthetic airport asset Excel workbook fixtures for focused tests."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

HEADERS = [
    "SL",
    "Device ID",
    "Building",
    "Level",
    "Zone",
    "DA",
    "System",
    "Area",
    "Location",
    "Device Type",
    "Day",
    "Keys Number",
    "Scissor Lift",
    "Mobile Scaffold",
    "Last Done",
    "Due Date",
    "Status",
    "Overdue",
    "Remarks",
]


def _row(
    sl: str,
    device_id: str,
    building: str = "TE3",
    level: str = "BAS",
    zone: str = "Z1",
    da: str = "DA21",
    system: str = "CCTV",
    area: str = "BASEMENT",
    location: str = "FIRE EXIT CORRIDOR",
    device_type: str = "CAMERA",
    day: str = "",
    keys: str = "",
    scaffold: str = "",
    mobile: str = "",
    last_done: str = "",
    due: str = "",
    status: str = "",
    overdue: str = "",
    remarks: str = "",
) -> list:
    return [
        sl, device_id, building, level, zone, da, system, area, location, device_type,
        day, keys, scaffold, mobile, last_done, due, status, overdue, remarks,
    ]


def build_zone1_rows() -> list[list]:
    rows = [
        _row("1", "TE3-CCT-BAS-DA21-FCT-001", system="CCT", location="FIRE EXIT CORRIDOR"),
        _row("2", "TE3-PAS-BAS-DA21-SPK-001", system="PAS", location="LOBBY A"),
        _row("3", "TE3-ACS-BAS-DA21-DRR-001", system="ACS"),
        _row("3B", "TE3-ACS-BAS-DA21-DRR-001", system="ACS"),
        _row("4", "TE3-RAS-BAS-DA21-SEN-001", system="RAS", level="GRD", zone="Z2"),
        _row("5", "TE3-TEL-BAS-DA31-PAN-001", system="TEL", da="DA31"),
        _row("7", "TE3-CCT-BAS-DA21-FCT-001", system="CCT", location="DIFFERENT LOCATION"),
        _row("8", "", location="MISSING DEVICE ID ROW"),
        _row("9", "TE3-CCT-BAS-DA21-FCT-002", location="", device_type=""),
        _row("10", "TE3-CCT-BAS-DA21-FCT-003", location="FIRE EXIT CORRIDOR DOOR"),
        _row("11", "TE3-CCT-BAS-DA21-FCT-004", day="FIRE EXIT CORRIDOR"),
        _row("12", "TE3-CCT-BAS-DA21-FCT-005", day="MONDAY", status="OK"),
    ]
    rows.append(_row("13", "TE3-CCT-BAS-DA21-FCT-006"))
    rows.extend([["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""] for _ in range(3)])
    return rows


def build_zone2_rows() -> list[list]:
    return [
        _row("201", "TE3-CCT-GRD-DA21-FCT-010", level="GRD", zone="Z2"),
        _row("202", "TE3-CCT-BAS-DA21-FCT-001", level="BAS", zone="Z1", location="CROSS SHEET DUPLICATE"),
        _row("203", "TE3-CCT-BAS-DA21-FCT-011", level="1ST"),
        _row("204", "TE3-CCT-2ND-DA31-FCT-012", level="2ND", da="DA31"),
    ]


def write_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for title, rows in (("Zone-1", build_zone1_rows()), ("Zone-2", build_zone2_rows())):
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_workbook_with_spaced_headers(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Zone-1")
    ws.append(["  device id ", "Building", "Level", "Zone", "DA", "System", "Area", "Location", "Device Type", "SL"])
    ws.append(["TE3-CCT-BAS-DA21-FCT-099", "TE3", "BAS", "Z1", "DA21", "CCT", "A", "LOC", "CAM", "99"])
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-CCT-GRD-DA21-FCT-100", level="GRD", zone="Z2"))
    wb.save(path)


def write_privacy_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet("Zone-1")
    ws.append(HEADERS)
    ws.append(_row("1", "TE3-CCT-BAS-DA21-FCT-P01", remarks="contact admin@example.com"))
    ws2 = wb.create_sheet("Zone-2")
    ws2.append(HEADERS)
    ws2.append(_row("1", "TE3-CCT-GRD-DA21-FCT-P02"))
    wb.save(path)
