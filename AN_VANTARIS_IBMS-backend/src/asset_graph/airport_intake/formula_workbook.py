"""Dual-view openpyxl workbook access for formula-safe airport intake."""
from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .constants import MAX_SOURCE_ROWS, REQUIRED_WORKSHEETS
from .errors import AirportIntakeError
from .formula_safety import is_formula_cell, summarize_formula_cell
from .headers import normalize_header


class FormulaSafeWorkbook:
    """Read workbook twice: formula metadata view and cached-value view."""

    def __init__(self, path: Path) -> None:
        self.path = path.resolve()
        self._formula_view = load_workbook(self.path, data_only=False, read_only=True)
        self._cached_view = load_workbook(self.path, data_only=True, read_only=True)

    def close(self) -> None:
        self._formula_view.close()
        self._cached_view.close()

    def __enter__(self) -> "FormulaSafeWorkbook":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    @staticmethod
    def validate_extension(path: Path) -> None:
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return
        if suffix in {".xls", ".xlsm", ".xlsb"}:
            raise AirportIntakeError("UNSUPPORTED_EXTENSION", f"unsupported workbook extension: {suffix}")
        raise AirportIntakeError("UNSUPPORTED_EXTENSION", "only .xlsx workbooks are accepted")

    @staticmethod
    def _validate_archive(path: Path) -> None:
        try:
            with zipfile.ZipFile(path, "r") as archive:
                names = archive.namelist()
        except zipfile.BadZipFile as exc:
            raise AirportIntakeError("INVALID_WORKBOOK", "workbook is not a valid xlsx archive") from exc
        if any(name.startswith("xl/vbaProject") for name in names):
            raise AirportIntakeError("MACROS_NOT_ALLOWED", "macro-enabled workbook content is not accepted")
        if any(name.startswith("xl/externalLinks/") for name in names):
            raise AirportIntakeError("EXTERNAL_LINKS", "external workbook links are not accepted")
        suspicious = ("xl/embeddings/", "xl/activeX/", "xl/customData/")
        if any(name.startswith(prefix) for name in names for prefix in suspicious):
            raise AirportIntakeError("HIDDEN_BINARY_PAYLOAD", "hidden binary payloads are not accepted")

    @staticmethod
    def open(path: Path) -> "FormulaSafeWorkbook":
        resolved = path.resolve()
        FormulaSafeWorkbook.validate_extension(resolved)
        if not resolved.is_file():
            raise AirportIntakeError("WORKBOOK_NOT_FOUND", "workbook file was not found")
        FormulaSafeWorkbook._validate_archive(resolved)
        book = FormulaSafeWorkbook(resolved)
        missing = [name for name in REQUIRED_WORKSHEETS if name not in book.sheet_names()]
        if missing:
            book.close()
            raise AirportIntakeError("MISSING_WORKSHEET", f"required worksheets missing: {', '.join(missing)}")
        return book

    def sheet_names(self) -> tuple[str, ...]:
        return tuple(self._formula_view.sheetnames)

    def rows_with_formula_metadata(self, sheet_name: str) -> list[tuple[int, dict[str, Any], dict[str, dict[str, Any]]]]:
        if sheet_name not in self._formula_view.sheetnames:
            raise AirportIntakeError("MISSING_WORKSHEET", f"worksheet not found: {sheet_name}")
        formula_ws = self._formula_view[sheet_name]
        cached_ws = self._cached_view[sheet_name]
        header_row = list(next(formula_ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        while header_row and (header_row[-1] is None or str(header_row[-1]).strip() == ""):
            header_row.pop()
        headers = [normalize_header("" if value is None else str(value)) for value in header_row]
        if any(not header for header in headers):
            raise AirportIntakeError("EMPTY_HEADER", f"worksheet {sheet_name} contains empty header cells")
        if len(headers) != len(set(headers)):
            raise AirportIntakeError("DUPLICATE_HEADERS", f"worksheet {sheet_name} contains duplicate headers")

        records: list[tuple[int, dict[str, Any], dict[str, dict[str, Any]]]] = []
        row_sequence = 0
        for row_index, (formula_row, cached_row) in enumerate(
            zip(
                formula_ws.iter_rows(min_row=2, values_only=True),
                cached_ws.iter_rows(min_row=2, values_only=True),
            ),
            start=2,
        ):
            record = {
                headers[col_index]: cached_row[col_index] if col_index < len(cached_row) else None
                for col_index in range(len(headers))
            }
            if not any(value not in (None, "") for value in record.values()):
                continue
            row_sequence += 1
            formula_meta: dict[str, dict[str, Any]] = {}
            location_value = "" if record.get("Location") is None else str(record.get("Location"))
            for col_index, header in enumerate(headers):
                if not header:
                    continue
                formula_value = formula_row[col_index] if col_index < len(formula_row) else None
                cached_value = cached_row[col_index] if col_index < len(cached_row) else None
                if is_formula_cell(formula_value):
                    formula_meta[header] = summarize_formula_cell(
                        worksheet=sheet_name,
                        column_name=header,
                        row_number=row_index,
                        formula_value=formula_value,
                        cached_value=cached_value,
                        location_value=location_value,
                    )
            records.append((row_index, record, formula_meta))
            if len(records) > MAX_SOURCE_ROWS:
                raise AirportIntakeError("EXCESSIVE_ROW_COUNT", f"worksheet {sheet_name} exceeds row policy")
        return records

    def formula_inventory(self) -> dict[str, Any]:
        inventory: dict[str, Any] = {"byWorksheet": {}, "formulaCellCount": 0, "arrayFormulaCellCount": 0}
        for sheet_name in REQUIRED_WORKSHEETS:
            ws = self._formula_view[sheet_name]
            count = 0
            array_count = 0
            columns: set[str] = set()
            header_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            while header_row and (header_row[-1] is None or str(header_row[-1]).strip() == ""):
                header_row.pop()
            headers = [normalize_header("" if value is None else str(value)) for value in header_row]
            for row in ws.iter_rows(min_row=2):
                for col_index, cell in enumerate(row):
                    if is_formula_cell(cell.value):
                        count += 1
                        if headers[col_index] if col_index < len(headers) else "":
                            columns.add(get_column_letter(col_index + 1))
                        from .formula_safety import is_array_formula

                        if is_array_formula(cell.value):
                            array_count += 1
            inventory["byWorksheet"][sheet_name] = count
            inventory["formulaCellCount"] += count
            inventory["arrayFormulaCellCount"] += array_count
            inventory.setdefault("formulaColumnsByWorksheet", {})[sheet_name] = sorted(columns)
        return inventory
