"""Asset import workbook parsing and quality validation.

This module is intentionally local and read-only. It parses uploaded Excel
workbooks, generates a quality report, and never writes Asset Registry data.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timezone
import hashlib
from io import BytesIO
from typing import Any, Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "Device ID",
    "Building",
    "Level",
    "Zone",
    "System",
    "Location",
    "Device Type",
]
BLOCKER_REQUIRED_COLUMNS = {"Device ID", "Building", "Level", "Zone", "System", "Device Type"}
ALLOWED_SYSTEMS = {"PA", "ACS", "CCTV", "TEL", "IPTV", "RAS", "MCS"}
LOCATION_HINTS = {
    "BASEMET": "BASEMENT",
    "MARKUP": "MAKEUP",
    "MEETERS-GREETERS": "MEETERS GREETERS",
    "DEPT": "DEPARTURE",
    " RM": " ROOM",
    "&": "AND",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def combined_source_hash(file_hashes: Iterable[str]) -> str:
    digest = hashlib.sha256()
    for file_hash in sorted(file_hashes):
        digest.update(file_hash.encode("utf-8"))
    return digest.hexdigest()


def _issue(severity: str, category: str, message: str, count: int, sample: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    return {
        "severity": severity,
        "category": category,
        "message": message,
        "count": int(count),
        "sample": sample or [],
    }


def _sample_rows(rows: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    return [
        {
            "source_file": row.get("__source_file", ""),
            "sheet": row.get("__sheet", ""),
            "row": row.get("__row", ""),
            "SL": row.get("SL", ""),
            "Device ID": row.get("Device ID", ""),
            "Building": row.get("Building", ""),
            "Level": row.get("Level", ""),
            "Zone": row.get("Zone", ""),
            "System": row.get("System", ""),
            "Location": row.get("Location", ""),
            "Device Type": row.get("Device Type", ""),
        }
        for row in rows[:limit]
    ]


def parse_workbook(source_file_name: str, payload: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    file_hash = sha256_bytes(payload)
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    sheet_profiles: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm(item) for item in rows[0]]
        if not any(headers):
            continue

        sheet_count = 0
        for row_index, row in enumerate(rows[1:], start=2):
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if header:
                    record[header] = norm(row[index] if index < len(row) else "")
            if not any(record.values()):
                continue
            record["__source_file"] = source_file_name
            record["__source_hash"] = file_hash
            record["__sheet"] = sheet_name
            record["__row"] = row_index
            records.append(record)
            sheet_count += 1
        sheet_profiles.append({"sheet_name": sheet_name, "records": sheet_count})

    profile = {
        "file_name": source_file_name,
        "file_hash_sha256": file_hash,
        "records": len(records),
        "sheet_count": len(sheet_profiles),
        "sheets": sheet_profiles,
    }
    return records, profile


def build_quality_report(
    *,
    import_batch_id: str,
    source_file_name: str,
    source_file_hash: str,
    uploaded_at: str,
    validated_at: str,
    source_file_profiles: list[dict[str, Any]],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    all_columns: set[str] = set()
    for record in records:
        all_columns.update(record.keys())

    issues: list[dict[str, Any]] = []
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in all_columns]
    if missing_columns:
        issues.append(
            _issue(
                "BLOCKER",
                "required_schema_missing",
                f"Required columns missing: {missing_columns}",
                len(missing_columns),
                [{"missing_column": column} for column in missing_columns],
            )
        )

    for column in REQUIRED_COLUMNS:
        missing = [record for record in records if record.get(column, "") == ""]
        if missing:
            severity = "BLOCKER" if column in BLOCKER_REQUIRED_COLUMNS else "MAJOR"
            issues.append(
                _issue(
                    severity,
                    "required_field",
                    f"Missing required field: {column}",
                    len(missing),
                    _sample_rows(missing),
                )
            )

    device_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        device_id = record.get("Device ID", "")
        if device_id:
            device_groups[device_id].append(record)

    duplicate_ids = {device_id: rows for device_id, rows in device_groups.items() if len(rows) > 1}
    if duplicate_ids:
        duplicate_sample: list[dict[str, Any]] = []
        for device_id, rows in list(duplicate_ids.items())[:10]:
            for row in rows[:3]:
                duplicate_sample.append(
                    {
                        "Device ID": device_id,
                        "source_file": row.get("__source_file", ""),
                        "sheet": row.get("__sheet", ""),
                        "row": row.get("__row", ""),
                        "System": row.get("System", ""),
                        "Location": row.get("Location", ""),
                        "Device Type": row.get("Device Type", ""),
                    }
                )
        issues.append(
            _issue("BLOCKER", "duplicate_device_id", "Duplicate Device ID detected", len(duplicate_ids), duplicate_sample)
        )

        conflict_count = 0
        conflict_samples: list[dict[str, Any]] = []
        for device_id, rows in duplicate_ids.items():
            conflicts: dict[str, list[str]] = {}
            for field in ["System", "Location", "Device Type", "Level", "Zone"]:
                values = sorted({row.get(field, "") for row in rows})
                if len(values) > 1:
                    conflicts[field] = values
            if conflicts:
                conflict_count += 1
                if len(conflict_samples) < 10:
                    conflict_samples.append({"Device ID": device_id, "conflicts": conflicts})
        if conflict_count:
            issues.append(
                _issue(
                    "BLOCKER",
                    "duplicate_device_id_conflict",
                    "Duplicate Device ID has conflicting System / Location / Device Type / Level / Zone",
                    conflict_count,
                    conflict_samples,
                )
            )

    unknown_systems = [
        record for record in records if record.get("System", "") and record.get("System", "") not in ALLOWED_SYSTEMS
    ]
    if unknown_systems:
        issues.append(
            _issue(
                "MAJOR",
                "system_not_allowed",
                "System value is not in approved dictionary",
                len(unknown_systems),
                _sample_rows(unknown_systems),
            )
        )

    locations = sorted({record.get("Location", "") for record in records if record.get("Location", "")})
    location_suggestions = []
    for location in locations:
        suggested = location
        for old, new in LOCATION_HINTS.items():
            suggested = suggested.replace(old, new)
        if suggested != location:
            location_suggestions.append({"location": location, "suggested": suggested})
    if location_suggestions:
        issues.append(
            _issue(
                "WARNING",
                "location_naming_normalization_suggestion",
                "Location naming normalization suggestions detected",
                len(location_suggestions),
                location_suggestions[:30],
            )
        )

    for column in ["Last Done", "Due Date", "Status", "Overdue"]:
        if column in all_columns:
            missing = [record for record in records if record.get(column, "") == ""]
            if missing:
                issues.append(
                    _issue(
                        "WARNING",
                        "maintenance_field_missing",
                        f"Missing maintenance field: {column}",
                        len(missing),
                        _sample_rows(missing, 5),
                    )
                )

    for column in ["Area", "Remarks"]:
        if column in all_columns:
            missing = [record for record in records if record.get(column, "") == ""]
            if missing:
                issues.append(_issue("WARNING", "optional_field_empty", f"Empty optional field: {column}", len(missing), []))

    observed = {
        "zones": sorted({record.get("Zone", "") for record in records if record.get("Zone", "")}),
        "da": sorted({record.get("DA", "") for record in records if record.get("DA", "")}),
        "locations": locations,
        "device_types": sorted({record.get("Device Type", "") for record in records if record.get("Device Type", "")}),
        "systems": sorted({record.get("System", "") for record in records if record.get("System", "")}),
    }

    severity_counts = Counter(issue["severity"] for issue in issues)
    readiness = "READY"
    if severity_counts["BLOCKER"] > 0:
        readiness = "HOLD_BLOCKED"
    elif severity_counts["MAJOR"] > 0:
        readiness = "REVIEW_REQUIRED"
    elif severity_counts["WARNING"] > 0:
        readiness = "READY_WITH_WARNINGS"

    confirm_enabled = readiness != "HOLD_BLOCKED"
    requires_second_confirmation = severity_counts["WARNING"] > 0 or severity_counts["MAJOR"] > 0
    summary = {
        "import_batch_id": import_batch_id,
        "source_file_name": source_file_name,
        "source_file_hash": source_file_hash,
        "uploaded_at": uploaded_at,
        "validated_at": validated_at,
        "total_records": len(records),
        "source_files": [profile["file_name"] for profile in source_file_profiles],
        "sheet_count": len({record.get("__sheet", "") for record in records if record.get("__sheet", "")}),
        "level_count": len({record.get("Level", "") for record in records if record.get("Level", "")}),
        "zone_count": len(observed["zones"]),
        "system_count": len(observed["systems"]),
        "device_type_count": len(observed["device_types"]),
        "location_count": len(locations),
        "duplicate_device_id_count": len(duplicate_ids),
        "missing_required_field_count": sum(
            issue["count"] for issue in issues if issue["category"] in {"required_schema_missing", "required_field"}
        ),
        "unmatched_location_count": 0,
        "blocker_count": severity_counts.get("BLOCKER", 0),
        "major_count": severity_counts.get("MAJOR", 0),
        "warning_count": severity_counts.get("WARNING", 0),
        "info_count": severity_counts.get("INFO", 0),
        "readiness": readiness,
        "confirm_enabled": confirm_enabled,
    }
    alert = {
        "title": "Asset Import Quality Check",
        "message": "Customer asset workbook has been parsed and validated. Review the quality report before import.",
        "readiness": readiness,
        "confirm_enabled": confirm_enabled,
        "requires_review": readiness in {"HOLD_BLOCKED", "REVIEW_REQUIRED", "READY_WITH_WARNINGS"},
        "requires_second_confirmation": requires_second_confirmation,
        "action_labels": [
            "Cancel Import",
            "Download Quality Report",
            "Review Issues",
            "Confirm Import",
        ],
    }
    return {
        "report_id": "ONE_ASSET_IMPORT_GA_R2B_QUALITY_REPORT",
        "generated_at": validated_at,
        "import_batch_id": import_batch_id,
        "source_file_name": source_file_name,
        "source_file_hash": source_file_hash,
        "source_file_profiles": source_file_profiles,
        "summary": summary,
        "alert": alert,
        "system_summary": Counter(record.get("System", "") for record in records).most_common(),
        "level_summary": Counter(record.get("Level", "") for record in records).most_common(),
        "zone_summary": Counter(record.get("Zone", "") for record in records).most_common(),
        "device_type_summary": Counter(record.get("Device Type", "") for record in records).most_common(80),
        "location_summary": Counter(record.get("Location", "") for record in records).most_common(100),
        "observed_dictionaries": observed,
        "issues": issues,
    }
